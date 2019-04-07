import openpyxl
import pickle
import os
from tqdm import tqdm
from openpyxl.styles import Font, Alignment, Side, Border

def main():
    results_conversion = {"title":2,"地址":3,"分类号":4,"申请号":5,"申请人":9,"专利权人":9,"发明人":10,"设计人":10,"申请日":12,"abstract":13,"申请公布日":15,"授权公告日":15}
    
    patent_class = 'publish'
    # patent_class = 'authorization'
    # patent_class = 'utility_model'
    # patent_class = 'design'

    excelfile='C:\\Files\\Documents\\apollo项目组\\国防科工局成果转化目录\\专利信息爬取_' + patent_class + '.xlsx'
    pklfile_step1 = 'results\\' + patent_class + '\\' + patent_class + '_step1.pkl'
    pklfile_step2 = 'results\\' + patent_class + '\\' + patent_class + '_step2.pkl'

    with open(pklfile_step2, 'rb') as f:
        results = pickle.load(f)

    # create a new excel file
    wb = openpyxl.Workbook()
    sheet = wb.get_sheet_by_name('Sheet')
    column = ('序号','发明名称','地址','分类号','申请号','专利类型','技术领域','应用领域','申请人/专利权人','发明人/设计人','法律状态','申请日','摘要/简要说明','转化方式','申请/授权公布日','解密公告日','发布时间','数据来源')

    # 加粗字体
    boldFont = Font(bold=True)
    # 上下、左右对齐，自动换行
    centerAlignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # 边框属性设置
    thin = Side(border_style="thin", color="000000")
    thinBorder = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 表格首行字体与对齐设置
    for columnNum in range(len(column)): # skip the first row
        sheet.cell(row=1, column=columnNum+1).value = column[columnNum]
        sheet.cell(row=1, column=columnNum+1).font = boldFont
        sheet.cell(row=1, column=columnNum+1).alignment = centerAlignment

    # 冻结行1
    sheet.freeze_panes = 'A2'

    # 从第二行开始写入数据
    init_row = 2
    for result in results:
        for num in range(result['page_size']):  # 考虑全部页数
        # for num in range(1):  # 只考虑第一页
            page = result['patent'][num+1]
            for patent in page:
                sheet.cell(row=init_row, column=1).value = init_row - 1  # 首列写入序号
                sheet.cell(row=init_row, column=1).alignment = centerAlignment
                sheet.row_dimensions[init_row].height = 300  # 设置行高
                for k,v in results_conversion.items():
                    try:
                        # 把pickle文件中'分类号'字符串里含'全部 专利代理机构'的部分删掉
                        if k == "分类号":
                            sheet.cell(row=init_row, column=v).value = patent[k].replace('全部 专利代理机构', '')
                        else:
                            sheet.cell(row=init_row, column=v).value = patent[k]
                        sheet.cell(row=init_row, column=v).alignment = centerAlignment
                    except KeyError as e:
                        pass
                init_row += 1
                if init_row % 10000 == 0:
                    print(f"已经写入了{init_row}行")

    # 设置单元格边框
    # for row in sheet.iter_rows(min_col=1, min_row=1, max_row=sheet.max_row, max_col=sheet.max_column):
    #     for cell in row:
    #         cell.border = thinBorder

    wb.save(excelfile)
    print(f"本文件共有{sheet.max_row-1}个专利。")
if __name__ == "__main__":
    main()
